from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from io import BytesIO
from django.db import transaction
from django.db.models import Avg, Count, Exists, Max, OuterRef, Sum
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from datetime import date, datetime, time, timedelta
from urllib.parse import urlencode
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from django.db.models import Q
from django.db.models.functions import TruncDay, TruncHour, TruncMonth
from access_control.decorators import verificar_permiso
from settings.context_processors import system_date_context

from .models import (
    CertificadoSII,
    TareaRPETC,
    TareaCesionRPETC,
    CesionRPETC,
    RevisionCesionRPETC,
    RevisionCesionComentario,
    LecturaAutomaticaConfig,
    LecturaAutomaticaEjecucion,
)
from access_control.models import Empresa
from access_control.models import Permiso, Vista
from .forms import CertificadoUploadForm
from .utils.maestro import get_maestroempresa_by_codigo

try:
    from auditoria.helpers import audit_log
except Exception:
    audit_log = None


def _normalizar_rut_filtro(value):
    raw = (value or '').strip().replace('.', '').replace(' ', '')
    if not raw:
        return None, None
    if '-' in raw:
        rut, dv = raw.rsplit('-', 1)
        return rut or None, dv.upper() or None
    return raw, None


def _fecha_sistema_request(request):
    value = system_date_context(request).get('fecha_sistema')
    try:
        return date.fromisoformat(value) if value else None
    except ValueError:
        return None


def _cesiones_rpetc_context(empresa_activa, fecha_seleccionada=None, filtros=None, include_detail=True):
    """Construye resumen y detalle sin duplicar cesiones por tareas repetidas."""
    if not empresa_activa:
        return {
            'cesiones_por_fecha': [],
            'fecha_cesion_seleccionada': None,
            'cesiones_detalle': [],
            'total_cesiones_rpetc': 0,
            'monto_total_cedido': 0,
            'ultima_cesion': None,
            'ultima_sincronizacion': None,
            'tipos_doc_filtro': [],
            'estados_filtro': [],
            'filtros': {},
            'filtros_querystring': '',
            'filtros_error': None,
            'proveedores_filtro': [],
        }

    filtros = filtros or {}
    base_cesiones = CesionRPETC.objects.filter(
        tareas__tarea__empresa=empresa_activa,
    ).distinct()
    tipos_doc = sorted(set(base_cesiones.values_list('tipo_doc', flat=True)))
    estados = sorted(set(base_cesiones.values_list('estado_cesion', flat=True)))
    proveedores = {}
    for rut, dv, razon in base_cesiones.values_list('cedente_rut', 'cedente_dv', 'cedente_razon_social'):
        if not rut:
            continue
        value = f'{rut}-{dv}' if dv else rut
        proveedores[value] = {
            'value': value,
            'label': f'{value} - {razon}' if razon else value,
            'razon_social': razon or '',
        }
    proveedores = sorted(proveedores.values(), key=lambda item: (item['razon_social'].lower(), item['value']))
    cesiones = base_cesiones
    proveedor_rut, proveedor_dv = _normalizar_rut_filtro(filtros.get('rut_proveedor'))
    if proveedor_rut:
        proveedor_q = Q(cedente_rut=proveedor_rut) | Q(vendedor_rut=proveedor_rut)
        if proveedor_dv:
            proveedor_q = (
                Q(cedente_rut=proveedor_rut, cedente_dv=proveedor_dv)
                | Q(vendedor_rut=proveedor_rut, vendedor_dv=proveedor_dv)
            )
        cesiones = cesiones.filter(proveedor_q)
    if filtros.get('tipo_doc'):
        cesiones = cesiones.filter(tipo_doc=filtros['tipo_doc'])
    if filtros.get('folio'):
        cesiones = cesiones.filter(folio_doc=filtros['folio'].strip())
    if filtros.get('estado'):
        cesiones = cesiones.filter(estado_cesion=filtros['estado'])
    if filtros.get('fecha_desde'):
        cesiones = cesiones.filter(fecha_cesion__date__gte=filtros['fecha_desde'])
    if filtros.get('fecha_hasta'):
        cesiones = cesiones.filter(fecha_cesion__date__lte=filtros['fecha_hasta'])
    summary = {}
    total_amount = 0
    for fecha, monto in cesiones.values_list('fecha_cesion', 'monto_cesion'):
        if not fecha:
            continue
        fecha_dia = timezone.localtime(fecha).date() if timezone.is_aware(fecha) else fecha.date()
        item = summary.setdefault(fecha_dia, {'fecha': fecha_dia, 'cantidad': 0, 'monto': 0})
        item['cantidad'] += 1
        item['monto'] += monto or 0
        total_amount += monto or 0
    summaries = sorted(summary.values(), key=lambda item: item['fecha'], reverse=True)[:30]
    detail = list(cesiones.order_by('-fecha_cesion', 'id_cesion')) if include_detail else []
    contabilidad = {}
    if detail:
        try:
            from .services.rpetc_contabilidad import obtener_estados_contables_cesiones
            contabilidad = obtener_estados_contables_cesiones(empresa_activa.codigo, detail)
        except Exception:
            contabilidad = {
                cesion.pk: {
                    'contabilizacion': {'estado': 'NO_DISPONIBLE', 'cantidad_movimientos': 0, 'movimientos': []},
                    'pagada_factoring': {'estado': 'NO_DISPONIBLE', 'cantidad_movimientos': 0, 'movimientos': []},
                    'pagada_proveedor': {'estado': 'NO_DISPONIBLE', 'cantidad_movimientos': 0, 'movimientos': []},
                }
                for cesion in detail
            }
    for cesion in detail:
        cesion.estado_contable = contabilidad.get(cesion.pk, {
            'contabilizacion': {'estado': 'NO_DISPONIBLE', 'cantidad_movimientos': 0, 'movimientos': []},
            'pago': {'estado': 'NO_DISPONIBLE', 'cantidad_movimientos': 0, 'movimientos': []},
        })
    latest = cesiones.order_by('-fecha_cesion').first()
    latest_sync = TareaRPETC.objects.filter(empresa=empresa_activa).order_by('-consultada_en').first()
    return {
        'cesiones_por_fecha': summaries,
        'fecha_cesion_seleccionada': fecha_seleccionada,
        'cesiones_detalle': detail,
        'total_cesiones_rpetc': cesiones.count(),
        'monto_total_cedido': total_amount,
        'ultima_cesion': latest,
        'ultima_sincronizacion': latest_sync,
        'tipos_doc_filtro': tipos_doc,
        'estados_filtro': estados,
        'filtros': filtros,
        'filtros_querystring': urlencode({
            key: value.isoformat() if isinstance(value, date) else value
            for key, value in filtros.items() if value
        }),
        'filtros_error': None,
        'proveedores_filtro': proveedores,
    }


def _rpetc_filtered_queryset(empresa_activa, filtros):
    base = CesionRPETC.objects.filter(
        tareas__tarea__empresa=empresa_activa,
    ).distinct()
    proveedor_rut, proveedor_dv = _normalizar_rut_filtro(filtros.get('rut_proveedor'))
    if proveedor_rut:
        if proveedor_dv:
            base = base.filter(
                Q(cedente_rut=proveedor_rut, cedente_dv=proveedor_dv)
                | Q(vendedor_rut=proveedor_rut, vendedor_dv=proveedor_dv)
            )
        else:
            base = base.filter(Q(cedente_rut=proveedor_rut) | Q(vendedor_rut=proveedor_rut))
    if filtros.get('tipo_doc'):
        base = base.filter(tipo_doc=filtros['tipo_doc'])
    if filtros.get('folio'):
        base = base.filter(folio_doc=filtros['folio'].strip())
    if filtros.get('estado'):
        base = base.filter(estado_cesion=filtros['estado'])
    if filtros.get('fecha_desde'):
        base = base.filter(fecha_cesion__date__gte=filtros['fecha_desde'])
    if filtros.get('fecha_hasta'):
        base = base.filter(fecha_cesion__date__lte=filtros['fecha_hasta'])
    return base


def _rpetc_annotated_queryset(empresa_activa, filtros):
    return _rpetc_filtered_queryset(empresa_activa, filtros).annotate(
        revision_registrada=Exists(RevisionCesionComentario.objects.filter(
            revision__empresa=empresa_activa, revision__cesion=OuterRef('pk'),
        )),
    )


def _rpetc_apply_payment_filters(filtered, empresa_activa, filtros):
    if not (filtros.get('sin_pago_factoring') or filtros.get('sin_pago_proveedor')):
        return filtered
    candidate_ids = list(filtered.values_list('pk', flat=True))
    states_by_pk = {}
    for start in range(0, len(candidate_ids), 250):
        chunk = list(CesionRPETC.objects.filter(pk__in=candidate_ids[start:start + 250]).order_by('pk'))
        if not chunk:
            continue
        from .services.rpetc_contabilidad import obtener_estados_contables_cesiones
        states_by_pk.update(obtener_estados_contables_cesiones(empresa_activa.codigo, chunk))
    pending_ids = _rpetc_pagos_pendientes_ids(
        states_by_pk,
        sin_pago_factoring=bool(filtros.get('sin_pago_factoring')),
        sin_pago_proveedor=bool(filtros.get('sin_pago_proveedor')),
    )
    return filtered.filter(pk__in=sorted(pending_ids))


def _rpetc_pagos_pendientes_ids(estados_por_pk, *, sin_pago_factoring=False, sin_pago_proveedor=False):
    """Return the ids that match the validated operational filters."""
    if not (sin_pago_factoring or sin_pago_proveedor):
        return set(estados_por_pk)

    pending_ids = set()
    for pk, state in (estados_por_pk or {}).items():
        state = state or {}
        factoring = state.get('pagada_factoring') or {}
        proveedor = state.get('pagada_proveedor') or {}
        factoring_estado = factoring.get('estado')
        proveedor_estado = proveedor.get('estado')

        factor_pending = factoring_estado in {'NO_PAGADA', 'REVISAR'} if factoring_estado else False
        proveedor_pending = proveedor_estado in {'NO_PAGADA', 'REVISAR'} if proveedor_estado else False

        if sin_pago_factoring and sin_pago_proveedor:
            if factor_pending and proveedor_pending:
                pending_ids.add(pk)
        elif sin_pago_factoring:
            if factor_pending:
                pending_ids.add(pk)
        elif sin_pago_proveedor:
            if proveedor_pending:
                pending_ids.add(pk)

    return pending_ids


def _rpetc_request_filters(request):
    today = _fecha_sistema_request(request)
    filters = {
        key: request.GET.get(key, '').strip()
        for key in ('rut_proveedor', 'tipo_doc', 'folio', 'estado', 'fecha_desde', 'fecha_hasta')
    }
    for key in ('sin_pago_factoring', 'sin_pago_proveedor', 'sin_revisar'):
        value = request.GET.get(key, '')
        filters[key] = str(value).strip().lower() in {'1', 'true', 'on', 'yes'}
    filters = {key: value for key, value in filters.items() if value not in ('', False, None)}
    if today:
        filters.setdefault('fecha_desde', date(today.year, 1, 1).isoformat())
        filters.setdefault('fecha_hasta', today.isoformat())
    for key in ('fecha_desde', 'fecha_hasta'):
        if key not in filters:
            continue
        try:
            filters[key] = date.fromisoformat(filters[key])
        except ValueError:
            filters.pop(key, None)
    return filters

@login_required
@verificar_permiso("Gestión DTE - Control de Cesiones", "ingresar")
def cesiones(request):
    """Vista principal: Control de Cesiones DTE."""
    # Datos de ejemplo/placeholder para la UI (KPIs)
    kpis = {
        'documentos_revisados': 0,
        'documentos_cedidos': 0,
        'documentos_no_cedidos': 0,
        'pendientes_revision': 0,
    }
    empresa_id = request.session.get('empresa_id')
    empresa_activa = Empresa.objects.filter(pk=empresa_id).first() if empresa_id else None
    from .forms import SincronizarCesionesRPETCForm
    can_sync = False
    if empresa_activa:
        vista = Vista.objects.filter(nombre='Gestión DTE - Control de Cesiones').first()
        permiso = Permiso.objects.filter(
            usuario=request.user, empresa=empresa_activa, vista=vista,
        ).first() if vista else None
        can_sync = bool(permiso and (permiso.modificar or permiso.supervisor))
    context = {
        'kpis': kpis,
        'empresa_activa': empresa_activa,
        'empresas_filtro': [empresa_activa] if empresa_activa else [],
        'can_sync_rpetc': can_sync,
        'sync_form': SincronizarCesionesRPETCForm(),
        'sync_result': None,
        'sync_error': None,
    }
    from .services.lectura_automatica import periodos_mensuales_rpetc
    fecha_sistema = _fecha_sistema_request(request)
    context['fecha_sistema_rpetc'] = fecha_sistema
    context['periodos_mensuales_rpetc'] = periodos_mensuales_rpetc(fecha_sistema) if fecha_sistema else []
    selected = request.GET.get('fecha_cesion')
    filtros = {
        key: request.GET.get(key, '').strip()
        for key in ('rut_proveedor', 'tipo_doc', 'folio', 'estado', 'fecha_desde', 'fecha_hasta')
    }
    for key in ('sin_pago_factoring', 'sin_pago_proveedor', 'sin_revisar'):
        value = request.GET.get(key, '')
        filtros[key] = str(value).strip().lower() in {'1', 'true', 'on', 'yes'}
    filtros = {key: value for key, value in filtros.items() if value not in ('', False, None)}
    today = _fecha_sistema_request(request)
    filtros_error = None
    if today:
        filtros.setdefault('fecha_desde', date(today.year, 1, 1))
        filtros.setdefault('fecha_hasta', today)
    else:
        filtros_error = 'No existe una fecha de sistema configurada.'
    for field in ('fecha_desde', 'fecha_hasta'):
        if field in filtros and isinstance(filtros[field], str):
            try:
                filtros[field] = date.fromisoformat(filtros[field])
            except ValueError:
                filtros_error = 'Las fechas del filtro no son válidas.'
                filtros.pop(field, None)
    if filtros.get('fecha_desde') and filtros.get('fecha_hasta') and filtros['fecha_desde'] > filtros['fecha_hasta']:
        filtros_error = 'La fecha desde no puede ser posterior a la fecha hasta.'
        filtros.pop('fecha_desde', None)
        filtros.pop('fecha_hasta', None)
    try:
        selected_date = date.fromisoformat(selected) if selected else None
    except ValueError:
        selected_date = None
    context.update(_cesiones_rpetc_context(empresa_activa, selected_date, filtros, include_detail=False))
    context['filtros_error'] = filtros_error
    return render(request, 'gestiondte/cesiones.html', context)


def _lectura_automatica_filas():
    from .services.lectura_automatica import empresas_elegibles

    elegibles = empresas_elegibles()
    codigos = [empresa.codigo for empresa, _certificado in elegibles]
    ejecuciones = LecturaAutomaticaEjecucion.objects.filter(
        empresa__codigo__in=codigos,
    ).select_related('empresa').order_by('empresa__codigo', '-ultima_actualizacion')
    latest = {}
    for ejecucion in ejecuciones:
        latest.setdefault(ejecucion.empresa.codigo, ejecucion)
    return [
        {
            'empresa': empresa,
            'certificado': certificado,
            'ejecucion': latest.get(empresa.codigo),
        }
        for empresa, certificado in elegibles
    ]


@login_required
@verificar_permiso("Gestión DTE - Lectura Automática de Cesiones", "ingresar")
def lectura_automatica_cesiones(request):
    from .services.lectura_automatica import rango_automatico

    config, _ = LecturaAutomaticaConfig.objects.get_or_create(pk=1)
    fecha_desde, fecha_hasta = rango_automatico()
    return render(request, 'gestiondte/lectura_automatica_cesiones.html', {
        'config': config,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'empresas_lectura': _lectura_automatica_filas(),
    })


@login_required
@verificar_permiso("Gestión DTE - Lectura Automática de Cesiones", "modificar")
def ejecutar_lectura_automatica_cesiones(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    from .services.lectura_automatica import (
        LecturaAutomaticaError,
        ejecutar_lote,
    )

    try:
        intervalo = int(request.POST.get('intervalo_minutos', 60))
        if intervalo not in {15, 30, 60}:
            raise LecturaAutomaticaError('Intervalo inválido.')
        config, _ = LecturaAutomaticaConfig.objects.get_or_create(pk=1)
        previous_config = {
            'habilitado': config.habilitado,
            'intervalo_minutos': config.intervalo_minutos,
        }
        config.habilitado = request.POST.get('habilitado') == 'on'
        config.intervalo_minutos = intervalo
        config.proxima_ejecucion = (
            timezone.now() + timedelta(minutes=intervalo)
            if config.habilitado else None
        )
        config.save(update_fields=['habilitado', 'intervalo_minutos', 'proxima_ejecucion', 'modificado'])
        if request.POST.get('action') == 'save':
            audit_log(
                request,
                'UPDATE',
                'gestiondte',
                object_type='configuracion_lectura_cesiones',
                object_id=str(config.pk),
                vista_nombre='Gestión DTE - Lectura Automática de Cesiones',
                before=previous_config,
                after={
                    'habilitado': config.habilitado,
                    'intervalo_minutos': config.intervalo_minutos,
                },
            )
            messages.success(request, 'Configuración guardada.')
            return redirect('gestion_dte:lectura_automatica_cesiones')
        fecha_desde = date.fromisoformat(request.POST.get('fecha_desde', ''))
        fecha_hasta = date.fromisoformat(request.POST.get('fecha_hasta', ''))
        resultado = ejecutar_lote(fecha_desde, fecha_hasta, tipo_ejecucion='MANUAL')
        audit_log(
            request,
            'EXECUTE',
            'gestiondte',
            object_type='lectura_cesiones',
            object_id=str(resultado.get('lote_id') or ''),
            vista_nombre='Gestión DTE - Lectura Automática de Cesiones',
            meta={
                'empresas_procesadas': len(resultado.get('ejecuciones') or []),
                'bloqueado': resultado.get('bloqueado', False),
            },
        )
        if resultado['bloqueado']:
            messages.warning(request, 'Ya existe una lectura de cesiones en proceso.')
        else:
            messages.success(request, 'Lectura de cesiones ejecutada.')
    except (ValueError, LecturaAutomaticaError) as exc:
        messages.error(request, str(exc))
    return redirect('gestion_dte:lectura_automatica_cesiones')


@login_required
@verificar_permiso("Gestión DTE - Lectura Automática de Cesiones", "ingresar")
def estado_lectura_automatica_cesiones(request):
    rows = _lectura_automatica_filas()
    return JsonResponse({
        'data': [
            {
                'empresa_codigo': row['empresa'].codigo,
                'empresa_nombre': row['empresa'].descripcion,
                'certificado': row['certificado'].estado_vigencia,
                'estado': row['ejecucion'].estado if row['ejecucion'] else 'PENDIENTE',
                'progreso': row['ejecucion'].progreso if row['ejecucion'] else 0,
                'ultima_actualizacion': row['ejecucion'].ultima_actualizacion.isoformat() if row['ejecucion'] else None,
            }
            for row in rows
        ],
    })


@login_required
@verificar_permiso("Gestión DTE - Control de Cesiones", "ingresar")
def cesiones_data(request):
    """Return one paginated, company-scoped DataTables page."""
    empresa_id = request.session.get('empresa_id')
    empresa_activa = Empresa.objects.filter(pk=empresa_id).first() if empresa_id else None
    if not empresa_activa:
        return JsonResponse({'draw': 0, 'recordsTotal': 0, 'recordsFiltered': 0, 'data': []}, status=403)

    filters = _rpetc_request_filters(request)
    base = CesionRPETC.objects.filter(tareas__tarea__empresa=empresa_activa).distinct()
    filtered = _rpetc_annotated_queryset(empresa_activa, filters)
    search = request.GET.get('search[value]', '').strip()
    if search:
        filtered = filtered.filter(
            Q(folio_doc__icontains=search)
            | Q(cedente_razon_social__icontains=search)
            | Q(cesionario_razon_social__icontains=search)
            | Q(cedente_rut__icontains=search)
            | Q(cesionario_rut__icontains=search)
        )
    sin_revisar = bool(filters.get('sin_revisar'))
    if sin_revisar:
        filtered = filtered.filter(revision_registrada=False)
    filtered = _rpetc_apply_payment_filters(filtered, empresa_activa, filters)
    total = base.values('pk').count()
    filtered_total = filtered.values('pk').count()
    filtered_ids = filtered.values('pk')
    amount_total = CesionRPETC.objects.filter(pk__in=filtered_ids).aggregate(
        total=Sum('monto_cesion'),
    ).get('total') or 0
    ordering = {
        '0': 'fecha_cesion',
        '1': 'tipo_doc',
        '2': 'folio_doc',
        '3': 'cedente_razon_social',
        '4': 'cesionario_razon_social',
        '5': 'monto_cesion',
        '6': 'estado_cesion',
    }
    order_field = ordering.get(request.GET.get('order[0][column]'), 'fecha_cesion')
    if request.GET.get('order[0][dir]') == 'asc':
        order_by = order_field
    else:
        order_by = f'-{order_field}'
    try:
        start = max(0, int(request.GET.get('start', 0)))
        length = int(request.GET.get('length', 25))
    except (TypeError, ValueError):
        start, length = 0, 25
    length = 100 if length < 1 or length > 100 else length
    page = list(filtered.order_by(order_by, 'pk')[start:start + length])
    try:
        from .services.rpetc_contabilidad import obtener_estados_contables_cesiones
        states = obtener_estados_contables_cesiones(empresa_activa.codigo, page)
    except Exception:
        states = {
            cesion.pk: {
                'contabilizacion': {'estado': 'NO_DISPONIBLE', 'movimientos': []},
                'pagada_factoring': {'estado': 'NO_DISPONIBLE', 'movimientos': []},
                'pagada_proveedor': {'estado': 'NO_DISPONIBLE', 'movimientos': []},
            }
            for cesion in page
        }

    def rut_display(rut, dv):
        return f'{rut}-{dv}' if rut and dv else rut or '-'

    def row_data(cesion):
        state = states.get(cesion.pk, {})
        return {
            'id': cesion.pk,
            'fecha_cesion': cesion.fecha_cesion.strftime('%d/%m/%Y %H:%M') if cesion.fecha_cesion else '-',
            'fecha_order': cesion.fecha_cesion.isoformat() if cesion.fecha_cesion else '',
            'tipo_doc': cesion.tipo_doc,
            'folio_doc': cesion.folio_doc,
            'cedente_nombre': cesion.cedente_razon_social or '-',
            'cedente_rut': rut_display(cesion.cedente_rut, cesion.cedente_dv),
            'cesionario_nombre': cesion.cesionario_razon_social or '-',
            'cesionario_rut': rut_display(cesion.cesionario_rut, cesion.cesionario_dv),
            'monto_cesion': str(cesion.monto_cesion or 0),
            'estado_cesion': cesion.estado_cesion,
            'revision': {'revisado': bool(getattr(cesion, 'revision_registrada', False))},
            'contabilizacion': {key: value for key, value in state.get('contabilizacion', {}).items() if key != 'movimientos'},
            'pagada_factoring': {key: value for key, value in state.get('pagada_factoring', state.get('pago', {})).items() if key != 'movimientos'},
            'pagada_proveedor': {key: value for key, value in state.get('pagada_proveedor', {}).items() if key != 'movimientos'},
        }

    try:
        draw = int(request.GET.get('draw', 0))
    except (TypeError, ValueError):
        draw = 0
    return JsonResponse({
        'draw': draw,
        'recordsTotal': total,
        'recordsFiltered': filtered_total,
        'summary': {'cantidad': filtered_total, 'monto_total': str(amount_total)},
        'data': [row_data(cesion) for cesion in page],
    })


def _excel_text(value):
    value = '' if value is None else str(value)
    return "'" + value if value[:1] in {'=', '+', '-', '@'} else value


def _excel_estado(value, payment=False):
    estados = {
        'CONTABILIZADA': 'Sí',
        'PAGADA': 'Sí',
        'PAGADA_FACTORING': 'Sí',
        'PAGADA_PROVEEDOR': 'Sí',
        'PAGADA_FACTORING_DIFERENCIA': 'Sí con diferencia',
        'PAGADA_PROVEEDOR_DIFERENCIA': 'Sí con diferencia',
        'NO_CONTABILIZADA': 'No',
        'NO_PAGADA': 'No',
        'REVISAR': 'Revisar',
        'TIPO_NO_SOPORTADO': 'No soportado',
        'NO_DISPONIBLE': 'No disponible',
    }
    return estados.get(value or '', value or '-')


@login_required
@verificar_permiso("Gestión DTE - Control de Cesiones", "ingresar")
def exportar_cesiones_excel(request):
    empresa_id = request.session.get('empresa_id')
    empresa_activa = Empresa.objects.filter(pk=empresa_id).first() if empresa_id else None
    if not empresa_activa:
        return JsonResponse({'error': 'Empresa activa requerida.'}, status=403)

    filters = _rpetc_request_filters(request)
    filtered = _rpetc_annotated_queryset(empresa_activa, filters)
    search = request.GET.get('search[value]', '').strip()
    if search:
        filtered = filtered.filter(
            Q(folio_doc__icontains=search)
            | Q(cedente_razon_social__icontains=search)
            | Q(cesionario_razon_social__icontains=search)
            | Q(cedente_rut__icontains=search)
            | Q(cesionario_rut__icontains=search)
        )
    if filters.get('sin_revisar'):
        filtered = filtered.filter(revision_registrada=False)
    filtered = _rpetc_apply_payment_filters(filtered, empresa_activa, filters)
    cesiones = list(filtered.order_by('-fecha_cesion', 'pk'))

    states = {}
    from .services.rpetc_contabilidad import obtener_estados_contables_cesiones
    for start in range(0, len(cesiones), 250):
        chunk = cesiones[start:start + 250]
        states.update(obtener_estados_contables_cesiones(empresa_activa.codigo, chunk))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Control de Cesiones'
    headers = [
        'Fecha cesión', 'Proveedor', 'RUT proveedor', 'Tipo documento', 'Folio',
        'Monto documento', 'Monto cesión', 'Cesionario', 'RUT cesionario',
        'Contabilizada', 'Pagada a factoring', 'Pagada a proveedor', 'Revisado',
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for cesion in cesiones:
        state = states.get(cesion.pk, {})
        fecha = cesion.fecha_cesion
        if fecha and timezone.is_aware(fecha):
            fecha = timezone.localtime(fecha).replace(tzinfo=None)
        sheet.append([
            fecha,
            _excel_text(cesion.cedente_razon_social),
            _excel_text(f'{cesion.cedente_rut}-{cesion.cedente_dv}' if cesion.cedente_rut and cesion.cedente_dv else cesion.cedente_rut),
            _excel_text(cesion.tipo_doc),
            _excel_text(cesion.folio_doc),
            cesion.monto_total,
            cesion.monto_cesion,
            _excel_text(cesion.cesionario_razon_social),
            _excel_text(f'{cesion.cesionario_rut}-{cesion.cesionario_dv}' if cesion.cesionario_rut and cesion.cesionario_dv else cesion.cesionario_rut),
            _excel_estado((state.get('contabilizacion') or {}).get('estado')),
            _excel_estado((state.get('pagada_factoring') or {}).get('estado'), payment=True),
            _excel_estado((state.get('pagada_proveedor') or {}).get('estado'), payment=True),
            'Sí' if getattr(cesion, 'revision_registrada', False) else 'No',
        ])
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions
    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(max(len(header) + 2, 14), 28)
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=7):
        if row[0].value:
            row[0].number_format = 'DD-MM-YYYY HH:MM'
        for cell in row[5:7]:
            cell.number_format = '#,##0'

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="control_cesiones_{empresa_activa.codigo}_{timezone.localdate().isoformat()}.xlsx"'
    return response


@login_required
@verificar_permiso("Gestión DTE - Control de Cesiones", "ingresar")
def detalle_contable_cesion(request, pk):
    empresa_id = request.session.get('empresa_id')
    empresa_activa = Empresa.objects.filter(pk=empresa_id).first() if empresa_id else None
    if not empresa_activa:
        return JsonResponse({'success': False, 'error': 'No hay empresa activa.'}, status=403)
    cesion = get_object_or_404(CesionRPETC.objects.filter(
        tareas__tarea__empresa=empresa_activa,
    ).distinct(), pk=pk)
    try:
        from .services.rpetc_contabilidad import obtener_detalle_contable_cesion
        detalle = obtener_detalle_contable_cesion(empresa_activa.codigo, cesion)
    except Exception:
        return JsonResponse({'success': False, 'error': 'No fue posible consultar el detalle contable.'}, status=503)

    def serializar(value):
        if hasattr(value, 'isoformat'):
            return value.isoformat()
        if hasattr(value, 'as_tuple'):
            return str(value)
        return value

    return JsonResponse({
        'success': True,
        'factura': {'tipo': cesion.tipo_doc, 'folio': cesion.folio_doc},
        'pagos_resumen': {
            role: {
                key: serializar(detalle.get(role, {}).get(key))
                for key in ('estado', 'monto_rpetc', 'monto_legacy', 'diferencia_monto')
                if key in detalle.get(role, {})
            }
            for role in ('pagada_factoring', 'pagada_proveedor')
        },
        'contabilizacion': [
            {key: serializar(value) for key, value in movimiento.items()}
            for movimiento in detalle['contabilizacion'].get('movimientos', [])
        ],
        'pago': [
            {key: serializar(value) for key, value in movimiento.items()}
            for movimiento in detalle.get('pagada_factoring', detalle.get('pago', {})).get('movimientos', [])
        ],
        'pagada_factoring': [
            {key: serializar(value) for key, value in movimiento.items()}
            for movimiento in detalle.get('pagada_factoring', {}).get('movimientos', [])
        ],
        'pagada_proveedor': [
            {key: serializar(value) for key, value in movimiento.items()}
            for movimiento in detalle.get('pagada_proveedor', {}).get('movimientos', [])
        ],
    })


def _revision_empresa(request):
    empresa_id = request.session.get('empresa_id')
    return Empresa.objects.filter(pk=empresa_id).first() if empresa_id else None


def _revision_cesion(request, pk, empresa):
    return get_object_or_404(CesionRPETC.objects.filter(
        tareas__tarea__empresa=empresa,
    ).distinct(), pk=pk)


def _revision_permisos(request, empresa):
    vista = Vista.objects.filter(nombre='Gestión DTE - Control de Cesiones').first()
    permiso = Permiso.objects.filter(usuario=request.user, empresa=empresa, vista=vista).first() if vista else None
    return {
        'crear': bool(permiso and (permiso.crear or permiso.supervisor)),
        'modificar': bool(permiso and (permiso.modificar or permiso.supervisor)),
        'eliminar': bool(permiso and (permiso.eliminar or permiso.supervisor)),
    }


def _revision_json(revision, permisos):
    usuario_id = permisos.get('_usuario_id')
    public_permisos = {key: permisos[key] for key in ('crear', 'modificar', 'eliminar')}
    if not revision:
        return {'revisado': False, 'comentarios': [], 'puede_agregar': public_permisos['crear'], 'permisos': public_permisos}
    comentarios = []
    for comentario in revision.comentarios.all():
        es_autor = comentario.creado_por_id == usuario_id
        comentarios.append({
            'id': comentario.pk,
            'glosa': comentario.comentario,
            'autor': comentario.creado_por.username if comentario.creado_por else None,
            'creado_en': comentario.creado_en.isoformat(),
            'modificado_en': comentario.modificado_en.isoformat(),
            'puede_editar': es_autor and public_permisos['modificar'],
            'puede_eliminar': es_autor and public_permisos['eliminar'],
        })
    return {
        'revisado': bool(comentarios),
        'comentarios': comentarios,
        'revision': {
            'glosa': comentarios[0]['glosa'] if comentarios else revision.glosa,
            'creado_por': comentarios[0]['autor'] if comentarios else None,
            'creado_en': comentarios[0]['creado_en'] if comentarios else revision.creado_en.isoformat(),
        },
        'puede_agregar': public_permisos['crear'],
        'permisos': public_permisos,
    }


@login_required
@verificar_permiso("Gestión DTE - Control de Cesiones", "ingresar")
def revision_cesion(request, pk):
    empresa = _revision_empresa(request)
    if not empresa:
        return JsonResponse({'error': 'Empresa activa requerida.'}, status=403)
    cesion = _revision_cesion(request, pk, empresa)
    revision = RevisionCesionRPETC.objects.filter(empresa=empresa, cesion=cesion).prefetch_related(
        'comentarios__creado_por',
    ).first()
    permisos = _revision_permisos(request, empresa)
    permisos['_usuario_id'] = request.user.pk
    return JsonResponse(_revision_json(revision, permisos))


def _validar_comentario_revision(request):
    comentario = (request.POST.get('comentario') or '').strip()
    if not comentario:
        raise ValueError('El comentario es obligatorio.')
    if len(comentario) > 2000:
        raise ValueError('El comentario no puede superar 2000 caracteres.')
    return comentario


def _revision_con_permisos(request, pk):
    empresa = _revision_empresa(request)
    if not empresa:
        return None, None, None
    cesion = _revision_cesion(request, pk, empresa)
    revision = RevisionCesionRPETC.objects.filter(empresa=empresa, cesion=cesion).first()
    permisos = _revision_permisos(request, empresa)
    permisos['_usuario_id'] = request.user.pk
    return empresa, cesion, (revision, permisos)


@login_required
@verificar_permiso("Gestión DTE - Control de Cesiones", "crear")
def crear_comentario_revision(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)
    empresa, cesion, datos = _revision_con_permisos(request, pk)
    if not empresa:
        return JsonResponse({'error': 'Empresa activa requerida.'}, status=403)
    try:
        comentario = _validar_comentario_revision(request)
    except ValueError as exc:
        return JsonResponse({'error': str(exc)}, status=400)
    revision, _ = RevisionCesionRPETC.objects.get_or_create(
        empresa=empresa,
        cesion=cesion,
        defaults={'glosa': comentario, 'creado_por': request.user},
    )
    RevisionCesionComentario.objects.create(
        revision=revision,
        comentario=comentario,
        creado_por=request.user,
    )
    return JsonResponse(_revision_json(revision, datos[1]), status=201)


def _comentario_autorizado(request, pk, comentario_pk):
    empresa, cesion, datos = _revision_con_permisos(request, pk)
    if not empresa:
        return None, None, None, None
    revision, permisos = datos
    comentario = get_object_or_404(
        RevisionCesionComentario,
        pk=comentario_pk,
        revision__empresa=empresa,
        revision__cesion=cesion,
    )
    return empresa, revision, comentario, permisos


@login_required
@verificar_permiso("Gestión DTE - Control de Cesiones", "modificar")
def editar_comentario_revision(request, pk, comentario_pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)
    empresa, revision, comentario_obj, permisos = _comentario_autorizado(request, pk, comentario_pk)
    if not empresa:
        return JsonResponse({'error': 'Empresa activa requerida.'}, status=403)
    if comentario_obj.creado_por_id != request.user.pk:
        return JsonResponse({'error': 'Solo el autor puede editar este comentario.'}, status=403)
    try:
        comentario_obj.comentario = _validar_comentario_revision(request)
    except ValueError as exc:
        return JsonResponse({'error': str(exc)}, status=400)
    comentario_obj.save(update_fields=['comentario', 'modificado_en'])
    return JsonResponse(_revision_json(revision, permisos))


@login_required
@verificar_permiso("Gestión DTE - Control de Cesiones", "eliminar")
def eliminar_comentario_revision(request, pk, comentario_pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)
    empresa, revision, comentario_obj, permisos = _comentario_autorizado(request, pk, comentario_pk)
    if not empresa:
        return JsonResponse({'error': 'Empresa activa requerida.'}, status=403)
    if comentario_obj.creado_por_id != request.user.pk:
        return JsonResponse({'error': 'Solo el autor puede eliminar este comentario.'}, status=403)
    comentario_obj.delete()
    if not revision.comentarios.exists():
        revision.delete()
        return JsonResponse({'revisado': False, 'comentarios': [], 'puede_agregar': permisos['crear'], 'permisos': {key: permisos[key] for key in ('crear', 'modificar', 'eliminar')}})
    return JsonResponse(_revision_json(revision, permisos))


@login_required
@verificar_permiso("Gestión DTE - Control de Cesiones", "crear")
def crear_revision_cesion(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)
    empresa = _revision_empresa(request)
    if not empresa:
        return JsonResponse({'error': 'Empresa activa requerida.'}, status=403)
    cesion = _revision_cesion(request, pk, empresa)
    glosa = (request.POST.get('glosa') or '').strip()
    if not glosa:
        return JsonResponse({'error': 'La glosa es obligatoria.'}, status=400)
    if len(glosa) > 2000:
        return JsonResponse({'error': 'La glosa no puede superar 2000 caracteres.'}, status=400)
    revision, creada = RevisionCesionRPETC.objects.get_or_create(
        empresa=empresa,
        cesion=cesion,
        defaults={'glosa': glosa, 'creado_por': request.user},
    )
    if creada:
        RevisionCesionComentario.objects.create(
            revision=revision,
            comentario=glosa,
            creado_por=request.user,
        )
    if not creada:
        return JsonResponse({'error': 'La cesión ya tiene una revisión registrada.'}, status=409)
    return JsonResponse(_revision_json(revision, _revision_permisos(request, empresa)), status=201)


@login_required
@verificar_permiso("Gestión DTE - Control de Cesiones", "modificar")
def editar_revision_cesion(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)
    empresa = _revision_empresa(request)
    if not empresa:
        return JsonResponse({'error': 'Empresa activa requerida.'}, status=403)
    cesion = _revision_cesion(request, pk, empresa)
    revision = get_object_or_404(RevisionCesionRPETC, empresa=empresa, cesion=cesion)
    glosa = (request.POST.get('glosa') or '').strip()
    if not glosa:
        return JsonResponse({'error': 'La glosa es obligatoria.'}, status=400)
    if len(glosa) > 2000:
        return JsonResponse({'error': 'La glosa no puede superar 2000 caracteres.'}, status=400)
    revision.glosa = glosa
    revision.modificado_por = request.user
    revision.save(update_fields=['glosa', 'modificado_por', 'modificado_en'])
    return JsonResponse(_revision_json(revision, _revision_permisos(request, empresa)))


@login_required
@verificar_permiso("Gestión DTE - Control de Cesiones", "eliminar")
def eliminar_revision_cesion(request, pk):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido.'}, status=405)
    empresa = _revision_empresa(request)
    if not empresa:
        return JsonResponse({'error': 'Empresa activa requerida.'}, status=403)
    cesion = _revision_cesion(request, pk, empresa)
    revision = get_object_or_404(RevisionCesionRPETC, empresa=empresa, cesion=cesion)
    revision.delete()
    return JsonResponse({'revisado': False, 'revision': None, 'permisos': _revision_permisos(request, empresa)})


@login_required
@verificar_permiso("Gestión DTE - Control de Cesiones", "modificar")
def sincronizar_cesiones_rpetc(request):
    from django.utils import timezone
    from .forms import SincronizarCesionesRPETCForm
    from .services.rpetc import (
        RPETCAuthenticationError, RPETCError, RPETCRateLimitError,
        RPETCTaskFailedError, RPETCTaskTimeoutError, RPETCUnauthorizedError,
    )
    from .services.rpetc_importer import RPETCImportError
    from .services.lectura_automatica import (
        LecturaAutomaticaError,
        periodos_mensuales_rpetc,
        rango_mensual_rpetc,
        sincronizar_empresa_rpetc,
    )

    empresa_id = request.session.get('empresa_id')
    empresa_activa = Empresa.objects.filter(pk=empresa_id).first() if empresa_id else None
    if not empresa_activa:
        return redirect('access_control:seleccionar_empresa')

    fecha_sistema = _fecha_sistema_request(request)

    form = SincronizarCesionesRPETCForm(request.POST or None)
    context = {
        'kpis': {
            'documentos_revisados': 0,
            'documentos_cedidos': 0,
            'documentos_no_cedidos': 0,
            'pendientes_revision': 0,
        },
        'empresa_activa': empresa_activa,
        'empresas_filtro': [empresa_activa] if empresa_activa else [],
        'can_sync_rpetc': True,
        'sync_form': form,
        'sync_result': None,
        'sync_error': None,
        'fecha_sistema_rpetc': fecha_sistema,
        'periodos_mensuales_rpetc': periodos_mensuales_rpetc(fecha_sistema) if fecha_sistema else [],
    }
    context.update(_cesiones_rpetc_context(empresa_activa))
    if request.method != 'POST' or not form.is_valid():
        return render(request, 'gestiondte/cesiones.html', context)

    mes = request.POST.get('mes', '').strip()
    if mes:
        try:
            if not fecha_sistema:
                raise LecturaAutomaticaError('No existe una fecha de sistema configurada.')
            fecha_desde, fecha_hasta = rango_mensual_rpetc(fecha_sistema, int(mes))
        except (LecturaAutomaticaError, TypeError, ValueError) as exc:
            context['sync_error'] = str(exc)
            return render(request, 'gestiondte/cesiones.html', context)
    else:
        fecha_desde = form.cleaned_data['fecha_desde']
        fecha_hasta = form.cleaned_data['fecha_hasta']
        if not fecha_sistema or fecha_desde.year != fecha_sistema.year or fecha_hasta.year != fecha_sistema.year:
            context['sync_error'] = 'El período debe pertenecer al año de la fecha de sistema.'
            return render(request, 'gestiondte/cesiones.html', context)
        if fecha_hasta > fecha_sistema:
            context['sync_error'] = 'La fecha hasta no puede superar la fecha de sistema.'
            return render(request, 'gestiondte/cesiones.html', context)
    en_progreso = TareaRPETC.objects.filter(
        empresa=empresa_activa,
        tipo_consulta='DEUDOR',
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        estado__in=['CREADO', 'EN_PROCESO', 'EN PROCESO'],
    ).first()
    if en_progreso:
        context['sync_error'] = 'Ya existe una sincronización en progreso para ese período.'
        return render(request, 'gestiondte/cesiones.html', context)

    certificado = CertificadoSII.objects.filter(
        empresa_codigo=empresa_activa.codigo, activo=True,
    ).first()
    if not certificado:
        context['sync_error'] = 'No existe un certificado SII activo para la empresa activa.'
        return render(request, 'gestiondte/cesiones.html', context)

    try:
        maestro = get_maestroempresa_by_codigo(empresa_activa.codigo)
        sincronizado = sincronizar_empresa_rpetc(
            empresa_activa,
            fecha_desde,
            fecha_hasta,
            certificado=certificado,
            maestro=maestro,
            intervalo=3,
            max_intentos=20,
        )
        resultado = sincronizado['resultado']
        stats = sincronizado['stats']
        audit_log(
            request,
            'UPDATE',
            'gestiondte',
            object_type='cesiones_rpetc',
            object_id=str(resultado['tarea_inicial'].get('idTarea') or ''),
            vista_nombre='Gestión DTE - Control de Cesiones',
            meta={
                'procesados': stats.get('registros_recibidos', 0),
                'actualizados': stats.get('cesiones_actualizadas', 0),
            },
        )
        context['sync_result'] = {
            **{key: value for key, value in stats.items() if key != 'tarea'},
            'id_tarea': resultado['tarea_inicial'].get('idTarea'),
            'periodo': f'{fecha_desde:%d/%m/%Y} - {fecha_hasta:%d/%m/%Y}',
            'fecha': timezone.localtime(),
        }
        context.update(_cesiones_rpetc_context(empresa_activa))
    except LecturaAutomaticaError as exc:
        context['sync_error'] = str(exc)
    except RPETCImportError:
        context['sync_error'] = 'Los datos recibidos del SII no pudieron guardarse.'
    except RPETCError as exc:
        if isinstance(exc, RPETCRateLimitError):
            message = 'El SII limitó temporalmente las solicitudes. Intenta más tarde.'
        elif isinstance(exc, RPETCUnauthorizedError):
            message = 'La empresa activa no está autorizada para esta consulta.'
        elif isinstance(exc, RPETCAuthenticationError):
            message = 'No fue posible autenticar contra el SII.'
        elif isinstance(exc, RPETCTaskTimeoutError):
            message = 'La consulta al SII excedió el tiempo máximo de espera.'
        elif isinstance(exc, RPETCTaskFailedError):
            message = 'La tarea RPETC terminó con error.'
        else:
            message = 'No fue posible completar la sincronización RPETC.'
        context['sync_error'] = message
    return render(request, 'gestiondte/cesiones.html', context)


@login_required
@verificar_permiso("Gestión DTE - Control de Cesiones", "ingresar")
def verificar_cesion(request):
    """Vista para consulta individual de cesión (placeholder)."""
    return render(request, 'gestiondte/verificar.html')


def _dashboard_periodo(periodo):
    ahora = timezone.now()
    hoy = timezone.localtime(ahora).date()
    if periodo == 'hoy':
        inicio = hoy
    elif periodo == 'semana':
        inicio = hoy - timedelta(days=hoy.weekday())
    elif periodo == 'mes':
        inicio = hoy.replace(day=1)
    elif periodo == 'anio':
        inicio = hoy.replace(month=1, day=1)
    else:
        raise ValueError('Período inválido.')
    inicio_local = timezone.make_aware(datetime.combine(inicio, time.min))
    return inicio_local, ahora


def _dashboard_resumen(empresa_activa, periodo):
    inicio, ahora = _dashboard_periodo(periodo)
    cesiones = CesionRPETC.objects.filter(
        tareas__tarea__empresa=empresa_activa,
        fecha_cesion__gte=inicio,
        fecha_cesion__lte=ahora,
    ).distinct()
    aggregate = cesiones.aggregate(
        total_cesiones=Count('pk'),
        monto_total=Sum('monto_cesion'),
        cedentes=Count('cedente_rut', distinct=True),
        cesionarios=Count('cesionario_rut', distinct=True),
        monto_promedio=Avg('monto_cesion'),
        ultima_cesion=Max('fecha_cesion'),
    )
    if periodo == 'hoy':
        trunc = TruncHour('fecha_cesion', tzinfo=timezone.get_current_timezone())
        formato = '%H:00'
    elif periodo == 'anio':
        trunc = TruncMonth('fecha_cesion', tzinfo=timezone.get_current_timezone())
        formato = '%Y-%m'
    else:
        trunc = TruncDay('fecha_cesion', tzinfo=timezone.get_current_timezone())
        formato = '%d-%m'

    evolucion = list(cesiones.annotate(periodo=trunc).values('periodo').annotate(
        cantidad=Count('pk'), monto=Sum('monto_cesion'),
    ).order_by('periodo'))
    evolucion = [
        {
            'periodo': item['periodo'].strftime(formato),
            'cantidad': item['cantidad'],
            'monto': int(item['monto'] or 0),
        }
        for item in evolucion if item['periodo']
    ]
    estados = list(cesiones.values('estado_cesion').annotate(
        cantidad=Count('pk'), monto=Sum('monto_cesion'),
    ).order_by('-cantidad', 'estado_cesion'))
    top_cesionarios = list(cesiones.values(
        'cesionario_rut', 'cesionario_dv', 'cesionario_razon_social',
    ).annotate(monto=Sum('monto_cesion'), cantidad=Count('pk')).order_by('-monto')[:10])
    return {
        'periodo': periodo,
        'kpis': {
            'total_cesiones': aggregate['total_cesiones'] or 0,
            'monto_total': int(aggregate['monto_total'] or 0),
            'cedentes': aggregate['cedentes'] or 0,
            'cesionarios': aggregate['cesionarios'] or 0,
            'monto_promedio': int(aggregate['monto_promedio'] or 0),
            'ultima_cesion': aggregate['ultima_cesion'].isoformat() if aggregate['ultima_cesion'] else None,
        },
        'evolucion': evolucion,
        'estados_rpetc': estados,
        'actividad': evolucion,
        'cesionarios_top': top_cesionarios,
    }


@login_required
@verificar_permiso("Gestión DTE - Dashboard DTE-SII-RPETC", "ingresar")
def index(request):
    empresa_id = request.session.get('empresa_id')
    empresa_activa = Empresa.objects.filter(pk=empresa_id).first() if empresa_id else None
    return render(request, 'gestiondte/index.html', {
        'empresa_activa': empresa_activa,
        'periodo_dashboard': 'mes',
        'dashboard_kpis': [
            {'key': 'total_cesiones', 'label': 'Total cesiones', 'color': 'primary', 'icon': 'ri-file-text-line'},
            {'key': 'monto_total', 'label': 'Monto total cedido', 'color': 'success', 'icon': 'ri-money-dollar-circle-line'},
            {'key': 'cedentes', 'label': 'Cedentes distintos', 'color': 'info', 'icon': 'ri-user-3-line'},
            {'key': 'cesionarios', 'label': 'Cesionarios distintos', 'color': 'warning', 'icon': 'ri-group-line'},
            {'key': 'monto_promedio', 'label': 'Monto promedio', 'color': 'danger', 'icon': 'ri-bar-chart-line'},
            {'key': 'ultima_cesion', 'label': 'Última cesión', 'color': 'secondary', 'icon': 'ri-time-line'},
        ],
    })


@login_required
@verificar_permiso("Gestión DTE - Dashboard DTE-SII-RPETC", "ingresar")
def dashboard_resumen(request):
    periodo = request.GET.get('periodo', 'mes').strip().lower()
    if periodo not in {'hoy', 'semana', 'mes', 'anio'}:
        return JsonResponse({'error': 'Período inválido.'}, status=400)
    empresa_id = request.session.get('empresa_id')
    empresa_activa = Empresa.objects.filter(pk=empresa_id).first() if empresa_id else None
    if not empresa_activa:
        return JsonResponse({'error': 'No hay empresa activa.'}, status=302)
    return JsonResponse(_dashboard_resumen(empresa_activa, periodo))


@login_required
@verificar_permiso("Gestión DTE - Certificados PFX-DTE", "ingresar")
def certificados_list(request):
    empresa_id = request.session.get('empresa_id')
    active_empresa = Empresa.objects.filter(pk=empresa_id).first() if empresa_id else None
    certificados = CertificadoSII.objects.filter(
        empresa_codigo=active_empresa.codigo,
    ) if active_empresa else CertificadoSII.objects.none()
    # gather empresa info for codes present
    codigos = set(cert.empresa_codigo for cert in certificados)
    empresas_certificados = {c: get_maestroempresa_by_codigo(c) for c in codigos}
    # compute can_create for UI: if user has crear on active empresa OR on any empresa
    can_create_context = False
    can_delete_context = False
    vista = Vista.objects.filter(nombre="Gestión DTE - Certificados PFX-DTE").first()
    if vista:
        if empresa_id:
            permiso = Permiso.objects.filter(usuario=request.user, empresa_id=empresa_id, vista=vista).first()
            can_create_context = bool(permiso and getattr(permiso, 'crear', False))
            can_delete_context = bool(permiso and (getattr(permiso, 'eliminar', False) or getattr(permiso, 'supervisor', False)))
        else:
            # check any company where user has crear
            can_create_context = Permiso.objects.filter(usuario=request.user, vista=vista, crear=True).exists()

    return render(request, 'gestiondte/certificados.html', {
        'certificados': certificados,
        'empresas_certificados': empresas_certificados,
        'can_create_context': can_create_context,
        'can_delete_context': can_delete_context,
        'active_empresa_codigo': active_empresa.codigo if active_empresa else '',
    })


@login_required
@verificar_permiso("Gestión DTE - Certificados PFX-DTE", "eliminar")
def certificados_eliminar(request, pk):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)

    certificado = get_object_or_404(CertificadoSII, pk=pk)
    empresa_id = request.session.get('empresa_id')
    empresa_activa = Empresa.objects.filter(pk=empresa_id).first() if empresa_id else None
    if not empresa_activa or certificado.empresa_codigo != empresa_activa.codigo:
        return JsonResponse({'success': False, 'error': 'El certificado no pertenece a la empresa activa.'}, status=403)

    before = {
        'empresa_id': empresa_activa.id,
        'empresa_codigo': certificado.empresa_codigo,
        'archivo_presente': bool(certificado.archivo and certificado.archivo.name),
        'activo': certificado.activo,
        'titular': certificado.titular,
        'emisor_certificado': certificado.emisor_certificado,
        'numero_serie': certificado.numero_serie,
        'rut_titular': certificado.rut_titular,
    }
    archivo = certificado.archivo
    try:
        with transaction.atomic():
            certificado.delete()
        if archivo and archivo.name:
            archivo.delete(save=False)
        audit_log(
            request,
            'DELETE',
            'gestiondte',
            object_type='certificado_sii',
            object_id=str(pk),
            vista_nombre='Gestión DTE - Certificados PFX-DTE',
            before=before,
            meta={'empresa_id': empresa_activa.id, 'empresa_codigo': before['empresa_codigo'], 'archivo_presente': before['archivo_presente']},
        )
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': 'gestion_dte.certificados.delete.success'})
        return redirect('gestion_dte:certificados')
    except Exception:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'No se pudo eliminar el certificado.'}, status=400)
        return redirect('gestion_dte:certificados')


@login_required
@verificar_permiso("Gestión DTE - Certificados PFX-DTE", "crear")
def certificados_cargar(request, codigoempresa=None):
    if request.method == 'POST':
        form = CertificadoUploadForm(request.POST, request.FILES)
        if form.is_valid():
            # Validate PKCS#12 before saving file
            uploaded = request.FILES.get('archivo')
            password = form.cleaned_data.get('password') or ''
            try:
                from cryptography.hazmat.primitives.serialization.pkcs12 import load_key_and_certificates
                from cryptography.hazmat.primitives.serialization import BestAvailableEncryption
                from cryptography.x509 import Certificate
                pwd_bytes = password.encode() if password else None
                data = uploaded.read()
                try:
                    key, cert, additional = load_key_and_certificates(data, pwd_bytes)
                except TypeError:
                    # Some versions expect bytes password or None
                    key, cert, additional = load_key_and_certificates(data, pwd_bytes)
                if cert is None:
                    form.add_error('archivo', 'El archivo no contiene un certificado PKCS#12 válido')
                else:
                    # Extract metadata
                    from cryptography.x509.oid import NameOID
                    subject = cert.subject.rfc4514_string()
                    issuer = cert.issuer.rfc4514_string()
                    try:
                        cn = cert.subject.get_attributes_for_oid(NameOID.COMMON_NAME)[0].value
                    except Exception:
                        cn = subject
                    serial = str(cert.serial_number)
                    valido_desde = cert.not_valid_before
                    valido_hasta = cert.not_valid_after
                    # try to get RUT from serialNumber or CN
                    rut = None
                    try:
                        serial_attrs = cert.subject.get_attributes_for_oid(NameOID.SERIAL_NUMBER)
                        if serial_attrs:
                            rut = serial_attrs[0].value
                    except Exception:
                        rut = None

                    if not rut:
                        import re
                        m = re.search(r"(\d{7,8}-[\dKk])", cn)
                        if m:
                            rut = m.group(1)

                    # Create instance but avoid saving file until metadata set
                    instance = form.save(commit=False)
                    if request.user and not instance.pk:
                        instance.created_by = request.user
                    instance.updated_by = request.user
                    instance.titular = cn
                    instance.emisor_certificado = issuer
                    instance.numero_serie = serial
                    instance.rut_titular = rut
                    instance.valido_desde = valido_desde
                    instance.valido_hasta = valido_hasta
                    # finally save (this saves file to storage)
                    instance.save()
                    audit_log(
                        request,
                        'CREATE',
                        'gestiondte',
                        instance,
                        object_type='certificado_sii',
                        vista_nombre='Gestión DTE - Certificados PFX-DTE',
                    )
                    return redirect('gestion_dte:certificados')
            except ImportError:
                form.add_error(None, 'Biblioteca cryptography no disponible en el entorno')
            except Exception as e:
                form.add_error('archivo', f'Error al validar certificado: {e}')
    else:
        initial = {}
        if codigoempresa:
            initial['empresa_codigo'] = codigoempresa
        else:
            # preferir la empresa activa en sesión cuando exista
            ses_codigo = request.session.get('empresa_codigo')
            if ses_codigo:
                initial['empresa_codigo'] = ses_codigo
        form = CertificadoUploadForm(initial=initial)
    return render(request, 'gestiondte/certificados_cargar.html', {'form': form, 'codigoempresa': codigoempresa})


@login_required
@verificar_permiso("Gestión DTE - Certificados PFX-DTE", "ingresar")
def certificados_detail(request, codigoempresa):
    certificados = CertificadoSII.objects.filter(empresa_codigo=codigoempresa)
    empresa = get_maestroempresa_by_codigo(codigoempresa)
    return render(request, 'gestiondte/certificados_detail.html', {'certificados': certificados, 'codigoempresa': codigoempresa, 'empresa': empresa})


@login_required
@verificar_permiso("Gestión DTE - Certificados PFX-DTE", "modificar")
def certificados_toggle_active(request, pk):
    cert = get_object_or_404(CertificadoSII, pk=pk)
    before = {'activo': cert.activo}
    cert.activo = not cert.activo
    cert.updated_by = request.user
    cert.save()
    audit_log(
        request,
        'UPDATE',
        'gestiondte',
        cert,
        object_type='certificado_sii',
        vista_nombre='Gestión DTE - Certificados PFX-DTE',
        before=before,
        after={'activo': cert.activo},
    )
    return redirect('gestion_dte:certificados')


@login_required
@verificar_permiso("Gestión DTE - Certificados PFX-DTE", "modificar")
def certificados_probar_conexion(request, pk):
    """Prueba de autenticación contra el SII con el certificado indicado (solo auth, sin RPETC)."""
    from django.utils import timezone
    from .services.sii_auth import probar_autenticacion_sii, SiiAuthError
    from .utils.maestro import get_maestroempresa_by_codigo

    cert = get_object_or_404(CertificadoSII, pk=pk)
    empresa = get_maestroempresa_by_codigo(cert.empresa_codigo)

    resultado = {
        'cert': cert,
        'empresa': empresa,
        'fecha_prueba': timezone.now(),
        'success': False,
        'token_obtenido': False,
        'token_expira': None,
        'rut_envio_sii': None,
        'error': None,
        'http_status': None,
    }

    try:
        res = probar_autenticacion_sii(cert)
        # success se deriva exclusivamente de token_obtenido para evitar inconsistencias
        resultado.update({
            'success': res['success'] and res['token_obtenido'],
            'token_obtenido': res['token_obtenido'],
            'token_expira': res.get('token_expira'),
            'rut_envio_sii': res.get('rut_envio_sii'),
        })
        audit_log(
            request,
            'EXECUTE',
            'gestiondte',
            cert,
            object_type='certificado_sii',
            vista_nombre='Gestión DTE - Certificados PFX-DTE',
            meta={'result': 'success'},
        )
    except SiiAuthError as e:
        resultado['error'] = str(e)
        resultado['http_status'] = e.http_status
        audit_log(
            request,
            'EXECUTE',
            'gestiondte',
            cert,
            object_type='certificado_sii',
            vista_nombre='Gestión DTE - Certificados PFX-DTE',
            meta={'result': 'failure', 'http_status': e.http_status},
        )

    return render(request, 'gestiondte/certificados_probar.html', resultado)
