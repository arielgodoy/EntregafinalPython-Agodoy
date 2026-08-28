from datetime import date
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from access_control.models import Empresa, Permiso, Vista
from gestiondte.models import (
    CesionRPETC,
    RevisionCesionComentario,
    RevisionCesionRPETC,
    TareaCesionRPETC,
    TareaRPETC,
)


class ExportarCesionesExcelTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='excel-user', password='pass')
        self.empresa = Empresa.objects.create(codigo='09', descripcion='Empresa A')
        self.otra_empresa = Empresa.objects.create(codigo='10', descripcion='Empresa B')
        vista, _ = Vista.objects.get_or_create(nombre='Gestión DTE - Control de Cesiones')
        Permiso.objects.create(usuario=self.user, empresa=self.empresa, vista=vista, ingresar=True)
        self.client.force_login(self.user)
        self._set_empresa(self.empresa)
        self.tarea = TareaRPETC.objects.create(
            empresa=self.empresa, id_tarea='excel-task', tipo_consulta='DEUDOR',
            rut_consultado='1', dv_consultado='9', fecha_desde=date(2026, 1, 1),
            fecha_hasta=date(2026, 12, 31), formato='TXT', estado='TERMINADO',
        )

    def _set_empresa(self, empresa):
        session = self.client.session
        session['empresa_id'] = empresa.id
        session.save()

    def _create_cesion(self, index, *, empresa=None, folio=None, proveedor='Proveedor', amount=100):
        cesion = CesionRPETC.objects.create(
            id_cesion=f'excel-{empresa.codigo if empresa else "09"}-{index}',
            estado_cesion='Vigente', deudor_rut='1', deudor_dv='9', tipo_doc='33',
            folio_doc=folio or str(index), cedente_rut='76376142', cedente_dv='8',
            cedente_razon_social=proveedor, cesionario_razon_social='Cesionario',
            cesionario_rut='76682670', cesionario_dv='9', monto_total=Decimal(amount),
            monto_cesion=Decimal(amount), fecha_cesion=timezone.now(),
        )
        tarea = self.tarea
        if empresa and empresa != self.empresa:
            tarea = TareaRPETC.objects.create(
                empresa=empresa, id_tarea=f'excel-task-{index}', tipo_consulta='DEUDOR',
                rut_consultado='1', dv_consultado='9', fecha_desde=date(2026, 1, 1),
                fecha_hasta=date(2026, 12, 31), formato='TXT', estado='TERMINADO',
            )
        TareaCesionRPETC.objects.create(tarea=tarea, cesion=cesion, rol_consulta='DEUDOR')
        return cesion

    def _download(self, params=None, states=None):
        states = states or {}
        with patch('gestiondte.services.rpetc_contabilidad.obtener_estados_contables_cesiones', return_value=states):
            return self.client.get(reverse('gestion_dte:exportar_cesiones_excel'), params or {})

    def test_exporta_todas_las_filas_y_tipos_excel(self):
        cesiones = [self._create_cesion(index, amount=index) for index in range(1, 31)]
        states = {
            cesion.pk: {
                'contabilizacion': {'estado': 'CONTABILIZADA'},
                'pagada_factoring': {'estado': 'PAGADA_FACTORING'},
                'pagada_proveedor': {'estado': 'PAGADA_PROVEEDOR'},
            }
            for cesion in cesiones
        }
        response = self._download({'length': '25'}, states)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertIn('control_cesiones_09_', response['Content-Disposition'])
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        sheet = workbook['Control de Cesiones']
        self.assertEqual(sheet.max_row, 31)
        self.assertEqual(sheet.freeze_panes, 'A2')
        self.assertEqual(sheet.auto_filter.ref, 'A1:M31')
        self.assertEqual(sheet['A2'].data_type, 'd')
        self.assertEqual({sheet.cell(row=row, column=6).value for row in range(2, 32)}, set(range(1, 31)))
        self.assertEqual(sheet['F2'].data_type, 'n')
        self.assertEqual(sheet['J2'].value, 'Sí')

    def test_exporta_filtros_revision_y_formula_injection(self):
        sin_revision = self._create_cesion(1, folio='2383', proveedor='=SUM(A1:A2)')
        revisada = self._create_cesion(2, folio='2384')
        revision = RevisionCesionRPETC.objects.create(empresa=self.empresa, cesion=revisada, glosa='Revisada', creado_por=self.user)
        RevisionCesionComentario.objects.create(revision=revision, comentario='Revisada', creado_por=self.user)
        states = {cesion.pk: {
            'contabilizacion': {'estado': 'NO_CONTABILIZADA'},
            'pagada_factoring': {'estado': 'NO_PAGADA'},
            'pagada_proveedor': {'estado': 'NO_PAGADA'},
        } for cesion in (sin_revision, revisada)}
        response = self._download({'sin_revisar': '1', 'folio': '2383'}, states)
        sheet = load_workbook(BytesIO(response.content), data_only=False)['Control de Cesiones']
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet['B2'].value, "'=SUM(A1:A2)")
        self.assertEqual(sheet['M2'].value, 'No')

    def test_export_aisla_empresa_activa(self):
        own = self._create_cesion(1, folio='A')
        other = self._create_cesion(2, empresa=self.otra_empresa, folio='B')
        response = self._download(states={own.pk: {}, other.pk: {}})
        sheet = load_workbook(BytesIO(response.content), data_only=False)['Control de Cesiones']
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet['E2'].value, 'A')

    def test_usuario_sin_ingresar_es_denegado(self):
        Permiso.objects.filter(usuario=self.user, empresa=self.empresa).update(ingresar=False)
        response = self.client.get(reverse('gestion_dte:exportar_cesiones_excel'))
        self.assertEqual(response.status_code, 403)

    def test_boton_exportacion_esta_en_la_pantalla(self):
        response = self.client.get(reverse('gestion_dte:cesiones'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="exportarCesionesExcel"')
        self.assertContains(response, 'Exportar Excel')
        self.assertContains(response, "URLSearchParams(new FormData(document.getElementById('cesiones-filtros-form')))")

    def test_export_triple_and_usa_interseccion(self):
        rows = [self._create_cesion(index) for index in range(1, 5)]
        revision = RevisionCesionRPETC.objects.create(empresa=self.empresa, cesion=rows[1], glosa='Revisada', creado_por=self.user)
        RevisionCesionComentario.objects.create(revision=revision, comentario='Revisada', creado_por=self.user)
        states = {
            rows[0].pk: {'pagada_factoring': {'estado': 'NO_PAGADA'}, 'pagada_proveedor': {'estado': 'NO_PAGADA'}},
            rows[1].pk: {'pagada_factoring': {'estado': 'NO_PAGADA'}, 'pagada_proveedor': {'estado': 'NO_PAGADA'}},
            rows[2].pk: {'pagada_factoring': {'estado': 'PAGADA_FACTORING'}, 'pagada_proveedor': {'estado': 'NO_PAGADA'}},
            rows[3].pk: {'pagada_factoring': {'estado': 'NO_PAGADA'}, 'pagada_proveedor': {'estado': 'PAGADA_PROVEEDOR'}},
        }
        response = self._download({
            'sin_pago_factoring': '1', 'sin_pago_proveedor': '1', 'sin_revisar': '1',
        }, states)
        sheet = load_workbook(BytesIO(response.content), data_only=False)['Control de Cesiones']
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet['E2'].value, '1')
